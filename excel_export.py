"""
Module for exporting portfolio and analysis data to professionally formatted Excel files.
Uses openpyxl directly for fine-grained formatting control.
"""

from io import BytesIO
from typing import List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Theme colours
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL_A = PatternFill(start_color="f0f4f8", end_color="f0f4f8", fill_type="solid")
ALT_FILL_B = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="d0d0d0"),
    right=Side(style="thin", color="d0d0d0"),
    top=Side(style="thin", color="d0d0d0"),
    bottom=Side(style="thin", color="d0d0d0"),
)

# Number format strings
FMT_CURRENCY = '$#,##0.00'
FMT_PERCENT = '0.00%'
FMT_DATE = 'YYYY-MM-DD'
FMT_NUMBER = '#,##0.00'


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _apply_header_style(ws, num_cols: int) -> None:
    """Style the first row as a dark-blue header."""
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _apply_body_styles(
    ws,
    num_rows: int,
    num_cols: int,
    col_formats: Optional[dict] = None,
) -> None:
    """Apply alternating row fills, borders, and per-column number formats.

    Parameters
    ----------
    col_formats : dict mapping 1-based column index to an openpyxl number_format string.
    """
    if col_formats is None:
        col_formats = {}

    for row_idx in range(2, num_rows + 1):
        fill = ALT_FILL_A if row_idx % 2 == 0 else ALT_FILL_B
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if col_idx in col_formats:
                cell.number_format = col_formats[col_idx]


def _auto_column_width(ws, num_cols: int, num_rows: int, min_width: int = 10, max_width: int = 40) -> None:
    """Set each column width to fit the longest cell value."""
    for col_idx in range(1, num_cols + 1):
        best = min_width
        for row_idx in range(1, num_rows + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                length = len(str(val)) + 2
                if length > best:
                    best = length
        ws.column_dimensions[get_column_letter(col_idx)].width = min(best, max_width)


def _write_dataframe(
    ws,
    df: pd.DataFrame,
    headers: List[str],
    col_formats: Optional[dict] = None,
) -> None:
    """Write *headers* then *df* values into the worksheet and apply full styling."""
    # Headers
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Body
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Convert numpy/pandas types to native Python so openpyxl handles them
            if pd.isna(val):
                cell.value = None
            elif hasattr(val, "item"):
                cell.value = val.item()
            elif isinstance(val, pd.Timestamp):
                cell.value = val.to_pydatetime()
            else:
                cell.value = val

    num_rows = len(df) + 1  # header row counts
    num_cols = len(headers)
    _apply_header_style(ws, num_cols)
    _apply_body_styles(ws, num_rows, num_cols, col_formats)
    _auto_column_width(ws, num_cols, num_rows)


def _workbook_to_bytes(wb: Workbook) -> bytes:
    """Save workbook into an in-memory buffer and return raw bytes."""
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _col_format_map(headers: List[str], currency_cols: List[str],
                    percent_cols: List[str], date_cols: List[str]) -> dict:
    """Build a {1-based col index: format_string} dict from column name lists."""
    fmt = {}
    for idx, h in enumerate(headers, start=1):
        if h in currency_cols:
            fmt[idx] = FMT_CURRENCY
        elif h in percent_cols:
            fmt[idx] = FMT_PERCENT
        elif h in date_cols:
            fmt[idx] = FMT_DATE
    return fmt


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def export_portfolio(watchlist_df: pd.DataFrame, trades_df: pd.DataFrame, forex_df: pd.DataFrame) -> bytes:
    """Create a comprehensive Excel workbook with multiple sheets.

    Sheets
    ------
    1. **Cartera** - watchlist positions
    2. **Trades Acciones** - stock trades
    3. **Trades Forex** - forex trades
    4. **Resumen** - summary KPIs

    Each sheet has professional formatting (dark-blue header, alternating row
    colours, auto-width columns, currency / percentage / date formats).

    Returns
    -------
    bytes
        Raw bytes suitable for ``st.download_button(data=...)``.
    """
    wb = Workbook()

    # ---- Sheet 1: Cartera ------------------------------------------------
    ws_cartera = wb.active
    ws_cartera.title = "Cartera"

    cartera_headers = [
        "Ticker", "Nombre", "Sector", "Cantidad", "Precio Compra",
        "Precio Actual", "Valor Total", "Ganancia/Perdida", "Rentabilidad %",
        "Fecha Compra",
    ]
    # Build a sanitised copy keeping only the columns we can map
    cartera_df = _safe_subset(watchlist_df, len(cartera_headers))
    cartera_formats = _col_format_map(
        cartera_headers,
        currency_cols=["Precio Compra", "Precio Actual", "Valor Total", "Ganancia/Perdida"],
        percent_cols=["Rentabilidad %"],
        date_cols=["Fecha Compra"],
    )
    _write_dataframe(ws_cartera, cartera_df, cartera_headers, cartera_formats)

    # ---- Sheet 2: Trades Acciones ----------------------------------------
    ws_trades = wb.create_sheet("Trades Acciones")

    trades_headers = [
        "Fecha", "Ticker", "Tipo", "Cantidad", "Precio",
        "Comision", "Total", "Nota",
    ]
    trades_sub = _safe_subset(trades_df, len(trades_headers))
    trades_formats = _col_format_map(
        trades_headers,
        currency_cols=["Precio", "Comision", "Total"],
        percent_cols=[],
        date_cols=["Fecha"],
    )
    _write_dataframe(ws_trades, trades_sub, trades_headers, trades_formats)

    # ---- Sheet 3: Trades Forex -------------------------------------------
    ws_forex = wb.create_sheet("Trades Forex")

    forex_headers = [
        "Fecha", "Par", "Tipo", "Lotes", "Precio Entrada",
        "Precio Salida", "Ganancia/Perdida", "Pips", "Nota",
    ]
    forex_sub = _safe_subset(forex_df, len(forex_headers))
    forex_formats = _col_format_map(
        forex_headers,
        currency_cols=["Precio Entrada", "Precio Salida", "Ganancia/Perdida"],
        percent_cols=[],
        date_cols=["Fecha"],
    )
    _write_dataframe(ws_forex, forex_sub, forex_headers, forex_formats)

    # ---- Sheet 4: Resumen ------------------------------------------------
    ws_resumen = wb.create_sheet("Resumen")
    _write_summary_sheet(ws_resumen, watchlist_df, trades_df, forex_df)

    return _workbook_to_bytes(wb)


def export_analyses(analyses_df: pd.DataFrame) -> bytes:
    """Export stock analyses history to Excel.

    Sheets
    ------
    1. **Analisis** - all saved analyses
    2. **Evolucion** - grouped by ticker with delta columns

    Returns
    -------
    bytes
        Raw bytes suitable for ``st.download_button(data=...)``.
    """
    wb = Workbook()

    # ---- Sheet 1: Analisis -----------------------------------------------
    ws_analisis = wb.active
    ws_analisis.title = "Analisis"

    analisis_headers = [
        "Fecha", "Ticker", "Precio", "Recomendacion",
        "Precio Objetivo", "Potencial %", "Puntuacion", "Notas",
    ]
    analisis_sub = _safe_subset(analyses_df, len(analisis_headers))
    analisis_formats = _col_format_map(
        analisis_headers,
        currency_cols=["Precio", "Precio Objetivo"],
        percent_cols=["Potencial %"],
        date_cols=["Fecha"],
    )
    _write_dataframe(ws_analisis, analisis_sub, analisis_headers, analisis_formats)

    # ---- Sheet 2: Evolucion ----------------------------------------------
    ws_evol = wb.create_sheet("Evolucion")
    _write_evolution_sheet(ws_evol, analyses_df)

    return _workbook_to_bytes(wb)


# ---------------------------------------------------------------------------
# Private helpers for specific sheets
# ---------------------------------------------------------------------------

def _safe_subset(df: pd.DataFrame, expected_cols: int) -> pd.DataFrame:
    """Return *df* trimmed or padded so it has exactly *expected_cols* columns.

    If the incoming DataFrame has more columns than expected we take the first
    N; if fewer we pad with ``None`` columns.  This avoids index errors when
    the caller's DataFrame schema doesn't match exactly.
    """
    if df is None or df.empty:
        return pd.DataFrame([[None] * expected_cols])

    df = df.copy()
    current = len(df.columns)
    if current >= expected_cols:
        df = df.iloc[:, :expected_cols]
    else:
        for i in range(expected_cols - current):
            df[f"_pad_{i}"] = None
    return df


def _write_summary_sheet(
    ws,
    watchlist_df: pd.DataFrame,
    trades_df: pd.DataFrame,
    forex_df: pd.DataFrame,
) -> None:
    """Build the Resumen (summary / KPI) sheet."""
    kpis: List[tuple] = []

    # Portfolio KPIs
    if watchlist_df is not None and not watchlist_df.empty:
        num_positions = len(watchlist_df)
        # Attempt to compute total value from the 7th column (Valor Total)
        try:
            total_value = pd.to_numeric(watchlist_df.iloc[:, 6], errors="coerce").sum()
        except Exception:
            total_value = 0
        try:
            total_gain = pd.to_numeric(watchlist_df.iloc[:, 7], errors="coerce").sum()
        except Exception:
            total_gain = 0
        try:
            avg_return = pd.to_numeric(watchlist_df.iloc[:, 8], errors="coerce").mean()
        except Exception:
            avg_return = 0

        kpis.extend([
            ("Posiciones en Cartera", num_positions, None),
            ("Valor Total Cartera", total_value, FMT_CURRENCY),
            ("Ganancia/Perdida Total", total_gain, FMT_CURRENCY),
            ("Rentabilidad Promedio", avg_return, FMT_PERCENT),
        ])
    else:
        kpis.append(("Posiciones en Cartera", 0, None))

    # Trades KPIs
    if trades_df is not None and not trades_df.empty:
        kpis.append(("Total Trades Acciones", len(trades_df), None))
    else:
        kpis.append(("Total Trades Acciones", 0, None))

    if forex_df is not None and not forex_df.empty:
        kpis.append(("Total Trades Forex", len(forex_df), None))
        try:
            forex_gain = pd.to_numeric(forex_df.iloc[:, 6], errors="coerce").sum()
            kpis.append(("Ganancia/Perdida Forex", forex_gain, FMT_CURRENCY))
        except Exception:
            pass
    else:
        kpis.append(("Total Trades Forex", 0, None))

    # Write KPI table
    headers = ["Indicador", "Valor"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)

    col_formats: dict = {}
    for row_idx, (label, value, fmt) in enumerate(kpis, start=2):
        ws.cell(row=row_idx, column=1, value=label)
        cell = ws.cell(row=row_idx, column=2, value=value)
        if fmt:
            cell.number_format = fmt

    num_rows = len(kpis) + 1
    num_cols = 2
    _apply_header_style(ws, num_cols)
    _apply_body_styles(ws, num_rows, num_cols, col_formats)
    _auto_column_width(ws, num_cols, num_rows)


def _write_evolution_sheet(ws, analyses_df: pd.DataFrame) -> None:
    """Group analyses by ticker and add delta columns showing change over time."""
    headers = [
        "Ticker", "Fecha", "Precio", "Precio Objetivo",
        "Potencial %", "Puntuacion", "Delta Precio", "Delta Objetivo",
    ]

    if analyses_df is None or analyses_df.empty:
        _write_dataframe(ws, pd.DataFrame([[None] * len(headers)]), headers)
        return

    df = analyses_df.copy()

    # Ensure we have at least the required positional columns
    if len(df.columns) < 7:
        _write_dataframe(ws, pd.DataFrame([[None] * len(headers)]), headers)
        return

    # Rename to known names for easier handling
    cols = list(df.columns)
    work = pd.DataFrame()
    work["Ticker"] = df.iloc[:, 1]
    work["Fecha"] = df.iloc[:, 0]
    work["Precio"] = pd.to_numeric(df.iloc[:, 2], errors="coerce")
    work["Precio Objetivo"] = pd.to_numeric(df.iloc[:, 4], errors="coerce")
    work["Potencial %"] = pd.to_numeric(df.iloc[:, 5], errors="coerce")
    work["Puntuacion"] = pd.to_numeric(df.iloc[:, 6], errors="coerce")

    # Sort by ticker then date
    try:
        work["_sort_date"] = pd.to_datetime(work["Fecha"], errors="coerce")
        work = work.sort_values(["Ticker", "_sort_date"])
    except Exception:
        work = work.sort_values(["Ticker"])

    # Delta columns: difference from previous row within same ticker
    work["Delta Precio"] = work.groupby("Ticker")["Precio"].diff()
    work["Delta Objetivo"] = work.groupby("Ticker")["Precio Objetivo"].diff()

    # Drop helper column
    work = work.drop(columns=["_sort_date"], errors="ignore")

    evol_formats = _col_format_map(
        headers,
        currency_cols=["Precio", "Precio Objetivo", "Delta Precio", "Delta Objetivo"],
        percent_cols=["Potencial %"],
        date_cols=["Fecha"],
    )
    _write_dataframe(ws, work[headers], headers, evol_formats)
